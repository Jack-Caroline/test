#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# Filename:jinxiaojian
# Create Time: 2021/12/17 19:05

# D:\ChalseeUserS\cjcj0064\Desktop\data\syms.xlsx
# Sheet2         1 A : 编号   2 B ： 产品名称

# D:\ChalseeUserS\cjcj0064\Desktop\data\jy.xlsx
# 客户交易信息 (35)    5  E ： 编号   6  F ： 产品名称

# D:\ChalseeUserS\cjcj0064\Desktop\data\cp.xlsx
# 客户交易信息 (35)    5  E ： 编号   6  F ： 产品名称
import openpyxl


class Excel(object):

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sh = self.wb[self.sheet_name]

    def show_excel(self, row_num, column_num):
        """
        读取 excel 表格中数据
        :return:每读取一次，返回一个值
        """
        res = self.sh.cell(row=row_num, column=column_num)
        return res.value

    def update_excel(self, row_num, column_num, value_num):
        """
        修改 excel 表中字段
        :param row_num: 行
        :param column_num: 列
        :param value_num: 值
        :return:
        """
        self.sh.cell(row=row_num, column=column_num, value=value_num)

    def save_excel(self):
        """
        报错 excel 表格
        :return:
        """
        self.wb.save(self.file_name)

    def num_excel(self):
        """
        获取 excel 表格中一共多少行
        :return:返回行数
        """
        return self.sh.max_row

    # def __del__(self, file_name):
    #     self.file_name = file_name
    #     self.wb.save(self.file_name)   27-->28


if __name__ == '__main__':
    ex1 = Excel(r"D:\ChalseeUserS\cjcj0064\Desktop\data\交易数据导出.xlsx", "受让数据")
    ex2 = Excel(r"D:\ChalseeUserS\cjcj0064\Desktop\data\交易数据导出.xlsx", "转让数据")
    num1 = ex1.num_excel()
    num2 = ex2.num_excel()
    print("表1一共：{}条，表2一共：{}条".format(num1,num2))
    for i in range(2,num2+1):
        for j in range(2,num1+1):
            if ex2.show_excel(i,22) == ex1.show_excel(j,39):
                ex2.update_excel(i,23,"成功")
                break
    ex2.save_excel()
    print("结束！！")

# if __name__ == '__main__':
#     """给表格第一行添加序号"""
#     list=["Sheet0","缺少日期"]
#     for sheet in list:
#         ex1 = Excel(r"D:\ChalseeUserS\cjcj0064\Desktop\data\分红.xlsx", sheet)
#         num1 = ex1.num_excel()
#         print("客户表数据为{}条，用户数据为{}条".format(num1, num1))
#         # 给表格加序号
#         for a in range(2,num1+1):
#             ex1.update_excel(a,1,a-1)
#         ex1.save_excel()
#         print("{}表，序号录入完成".format(sheet))
#     print("全部录入完成")

# if __name__ == '__main__':
#     """比对两个excle的数据，不一致的标记 1"""
#     ex1 = Excel(r"D:\data\产品基本数据导出.xlsx", "Sheet1")
#     ex2 = Excel(r"D:\data\1.产品信息.xlsx", "Sheet2")
#     num1 = ex1.num_excel()
#     num2 = ex2.num_excel()
#     print("新表产品{}条，旧表产品{}条".format(num1, num2))
#     for i in range(2, num2 + 1):
#         for j in range(2, num1 + 1):
#             if ex2.show_excel(i, 2) == ex1.show_excel(j, 1):
#                 ex2.update_excel(i, 3, 1)
#     ex2.save_excel()
#     print("OK")

# if __name__ == '__main__':
#     """处理客户数据中，所属理财师，将理财师名称改为工号"""
#     ex1 = Excel(r"D:\ChalseeUserS\cjcj0064\Desktop\data\客户基本信息.xlsx", "Sheet1")
#     ex2 = Excel(r"D:\ChalseeUserS\cjcj0064\Desktop\data\2.融义用户信息.xlsx", "Sheet1")
#     num1 = ex1.num_excel()
#     num2 = ex2.num_excel()
#     print("客户表数据为{}条，用户数据为{}条".format(num1, num1))
#     # 给表格加序号
#     for a in range(2,num1+1):
#         ex1.update_excel(a,1,a-1)
#     print("序号录入完成")
#     for i in range(2, num2 + 1):
#         for j in range(2, num1 + 1):
#             if ex1.show_excel(j, 9) == ex2.show_excel(i, 3):
#                 value_name = str(ex2.show_excel(i, 2))
#                 ex1.update_excel(j, 9, value_num=value_name)
#             if ex1.show_excel(j,9) is None:
#                 ex1.update_excel(j, 9, value_num="admin.guakong")
#     ex1.save_excel()
#     print("数据更改完成")

# if __name__ == '__main__':
#     ex1 = Excel(r"D:\ChalseeUserS\cjcj0064\Desktop\data\交易数据导出.xlsx", "Sheet1")
#     ex2 = Excel(r"D:\ChalseeUserS\cjcj0064\Desktop\data\2.融义用户信息.xlsx", "Sheet1")
#     num1 = ex1.num_excel()
#     num2 = ex2.num_excel()
#     print("交易数据为{}条，用户数据为{}条".format(num1, num2))
#     # 给表格加序号
#     for a in range(2,num1+1):
#         ex1.update_excel(a,1,a-1)
#     print("序号录入完成")
#     # 循环替换，交易表中所属理财师，将汉字替换为工号
#     for i in range(2, num2 + 1):
#         for j in range(2, num1 + 1):
#             if ex1.show_excel(j, 32) == ex2.show_excel(i, 3):
#                 value_name = str(ex2.show_excel(i, 2))
#                 ex1.update_excel(j, 32, value_num=value_name)
#     ex1.save_excel()
#     print("完成")

# if __name__ == '__main__':
#     """修改客户表中日期格式，改为文本"""
#     ex1 = Excel(r"D:\ChalseeUserS\cjcj0064\Desktop\data\客户基本信息.xlsx", "Sheet1")
#     num = ex1.num_excel()
#     for i in range(2, num + 1):
#         if ex1.show_excel(i,26) is None:
#             continue
#         value_num = str(ex1.show_excel(i, 26))    # 3     26
#         ex1.update_excel(i, 27, value_num[0:10])  # 4     27
#     ex1.save_excel()

# if __name__ == '__main__':
#     """修改交易表中日期格式，改为文本"""
#     ex1 = Excel(r"D:\ChalseeUserS\cjcj0064\Desktop\data\交易数据导出.xlsx", "Sheet1")
#     num = ex1.num_excel()
#     for i in range(2, num + 1):
#         value_num = str(ex1.show_excel(i, 33))  # 12     27     33
#         ex1.update_excel(i, 34, value_num[0:10])  # 13     28    34
#     ex1.save_excel()


# if __name__ == '__main__':
#     ex1 = Excel(r'D:\ChalseeUserS\cjcj0064\Desktop\data\产品基本数据导出.xlsx', 'Sheet1')
#     ex2 = Excel(r'D:\ChalseeUserS\cjcj0064\Desktop\data\交易数据导出.xlsx', 'Sheet1')
#     num1 = ex1.num_excel()
#     num2 = ex2.num_excel()
#     print(num1)
#     print(num2)
#     print("开始运行")
#     for i in range(2, num1 + 1):
#         print(i)
#         for j in range(2, num2 + 1):
#             if ex2.show_excel(j, 6) == ex1.show_excel(i, 3):
#                 ex1_value = int(ex1.show_excel(i, 2))
#                 ex2.update_excel(j, 7, ex1_value)
#         ex2.save_excel()
#     print("运行结束")

# if __name__ == '__main__':
#     ex1 = Excel(r'D:\ChalseeUserS\cjcj0064\Desktop\zhuomian\1.产品信息.xlsx', '产品信息')
#     ex2 = Excel(r'D:\ChalseeUserS\cjcj0064\Desktop\zhuomian\4.交易记录.xlsx', '认申购')
#     num1 = ex1.num_excel()
#     num2 = ex2.num_excel()
#     print("开始执行")
#     for i in range(2, num1 + 1):
#         ex1_value_num = 0
#         for j in range(2, num2 + 1):
#             if str(ex1.show_excel(i, 2)) == str(ex2.show_excel(j, 7)):
#                 ex1_value_num += 1
#                 ex1.update_excel(i, 4, ex1_value_num)
#     ex1.wb.save(r'D:\ChalseeUserS\cjcj0064\Desktop\zhuomian\1.产品信息.xlsx')
#     print("执行结束")
